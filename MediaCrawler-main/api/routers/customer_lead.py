# -*- coding: utf-8 -*-
"""
获客线索API路由
"""
import io
import time
import uuid
import asyncio
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Query, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, model_validator
from sqlalchemy import select, desc, func
from database.db_session import get_session as async_db_session
from database.models import CustomerLead, CrawlerTaskModel
from ..services.auth import get_current_user, user_scope_filter, is_admin

router = APIRouter(prefix="/leads", tags=["customer_leads"])


# ==================== 批量任务状态注册表 ====================
# 用于跟踪"批量采集联系方式"和"批量监测回复"的进度,前端通过 job_id 轮询展示。
# 结构: {job_id: {job_id, task_id, kind, status, total, completed, success, failed, message, created_at, updated_at}}
_batch_jobs: Dict[str, Dict[str, Any]] = {}
_BATCH_JOB_TTL_SEC = 1800  # 30分钟后清理(避免内存无限增长)


def _register_batch_job(job_id: str, task_id: str, kind: str, total: int) -> Dict[str, Any]:
    """注册一个批量任务,返回任务状态字典(调用方持有引用并更新)。"""
    now_ms = int(time.time() * 1000)
    job = {
        "job_id": job_id,
        "task_id": task_id,
        "kind": kind,  # "contact_collect" | "reply_monitor"
        "status": "running",  # running | completed | failed
        "total": total,
        "completed": 0,
        "success": 0,
        "failed": 0,
        "message": "",
        "created_at": now_ms,
        "updated_at": now_ms,
    }
    _batch_jobs[job_id] = job
    _cleanup_stale_batch_jobs()
    return job


def _update_batch_job(job_id: str, **fields) -> None:
    """更新批量任务进度字段。"""
    job = _batch_jobs.get(job_id)
    if not job:
        return
    job.update(fields)
    job["updated_at"] = int(time.time() * 1000)


def _finish_batch_job(job_id: str, status: str, message: str = "") -> None:
    """标记批量任务结束(running -> completed/failed)。"""
    _update_batch_job(job_id, status=status, message=message)


def _cleanup_stale_batch_jobs() -> None:
    """清理超过 TTL 的旧任务(避免内存无限增长)。"""
    now = int(time.time() * 1000)
    stale = [jid for jid, j in _batch_jobs.items()
             if now - j.get("updated_at", 0) > _BATCH_JOB_TTL_SEC * 1000]
    for jid in stale:
        _batch_jobs.pop(jid, None)


class LeadResponse(BaseModel):
    id: int
    task_id: str
    platform: str
    data_type: str = ""
    data_id: str = ""
    user_id: str = ""
    sec_uid: str = ""
    nickname: str = ""
    avatar: str = ""
    ip_location: str = ""
    content: str = ""
    title: str = ""
    url: str = ""
    matched_keywords: str = ""
    intent_type: str = ""
    lead_score: int = 0
    status: str = ""
    notes: Optional[str] = None
    add_ts: int = 0
    last_modify_ts: int = 0
    create_time: Optional[int] = None
    # 来源视频/作品信息(用于营销时查看原视频)
    source_aweme_id: str = ""
    source_video_title: str = ""
    source_video_desc: str = ""
    source_video_url: str = ""
    source_cover_url: str = ""
    source_author_nickname: str = ""
    # 增强字段(客户需求:支持复制和打开链接)
    comment_url: str = ""
    profile_url: str = ""
    platform_display_id: str = ""
    # 去重字段:同一用户相似广告模板评论合并后的重复次数
    dup_count: int = 1
    # 角色标签:supplier供方/consumer需求方/neutral中性(行业无关行为信号分析,配合 target_role 过滤)
    role_tag: str = "neutral"
    # 用户主页联系方式(一键获客时从用户主页采集)
    contact_phone: str = ""
    contact_wechat: str = ""
    bio_text: str = ""
    # 联系方式采集状态 + 评论回复监测时间戳
    contact_status: str = "pending"
    reply_monitor_ts: int = 0

    @model_validator(mode="before")
    @classmethod
    def _normalize_null_fields(cls, data):
        """将数据库中可能为 NULL 的新增字段转为默认值,避免 Pydantic 验证失败。"""
        if hasattr(data, "__dict__"):
            # ORM 对象:将 None 值的字段转为默认值
            for field in ("role_tag", "contact_phone", "contact_wechat",
                          "bio_text", "contact_status"):
                if getattr(data, field, None) is None:
                    setattr(data, field, "")
            if getattr(data, "dup_count", None) is None:
                setattr(data, "dup_count", 1)
            if getattr(data, "reply_monitor_ts", None) is None:
                setattr(data, "reply_monitor_ts", 0)
        elif isinstance(data, dict):
            # 字典:将 None 值转为默认值
            defaults = {
                "role_tag": "neutral", "contact_phone": "", "contact_wechat": "",
                "bio_text": "", "contact_status": "pending",
                "dup_count": 1, "reply_monitor_ts": 0,
            }
            for k, v in defaults.items():
                if data.get(k) is None:
                    data[k] = v
        return data

    class Config:
        from_attributes = True


class LeadListResponse(BaseModel):
    total: int
    items: List[LeadResponse]
    page: int
    page_size: int


class LeadStatsResponse(BaseModel):
    total_leads: int
    new_leads: int
    pending_leads: int
    contacted_leads: int
    qualified_leads: int
    converted_leads: int
    failed_leads: int
    ignored_leads: int
    platform_distribution: dict
    intent_distribution: dict
    avg_lead_score: float
    role_distribution: dict


@router.get("/list", response_model=LeadListResponse)
async def list_leads(
    task_id: Optional[str] = None,
    platform: Optional[str] = None,
    intent_type: Optional[str] = None,
    status: Optional[str] = None,
    min_score: Optional[int] = None,
    max_score: Optional[int] = None,
    level: Optional[str] = None,
    keyword: Optional[str] = None,
    ip_location: Optional[str] = None,
    start_ts: Optional[int] = None,
    end_ts: Optional[int] = None,
    role_tag: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    current_user: dict = Depends(get_current_user),
):
    """获取获客线索列表(按用户隔离)

    Args:
        level: 意向等级筛选 high(>=80) / medium(50-79) / low(<50),与 min_score/max_score 互斥补充
        role_tag: 角色标签筛选 supplier供方/consumer需求方/neutral中性
    """
    import json
    async with async_db_session() as session:
        # 自动获取任务的 target_regions 并应用地区过滤
        auto_region_regions = None
        if task_id:
            try:
                from database.models import CrawlerTaskModel
                task_result = await session.execute(
                    select(CrawlerTaskModel.target_regions).where(CrawlerTaskModel.id == task_id)
                )
                task_row = task_result.first()
                if task_row and task_row[0]:
                    regions = json.loads(task_row[0])
                    if regions and isinstance(regions, list) and len(regions) > 0:
                        auto_region_regions = regions
            except Exception:
                pass

        query = select(CustomerLead)
        # 复用通用筛选(用户隔离 + 所有过滤条件,避免与 export/stats 出现不一致)
        query = _apply_lead_filters(
            query, task_id, platform, intent_type, status,
            min_score, max_score, level, keyword, ip_location,
            current_user, start_ts, end_ts, role_tag,
        )

        # 应用任务自动地区过滤(如果有设置 target_regions)
        if auto_region_regions and not ip_location:
            conditions = []
            for region in auto_region_regions:
                conditions.append(CustomerLead.ip_location.contains(region))
            # 用 OR 连接: ip_location 包含任一目标地区
            from sqlalchemy import or_
            query = query.where(or_(*conditions))

        # 获取总数
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar()

        # 分页查询
        query = query.order_by(desc(CustomerLead.lead_score), desc(CustomerLead.add_ts))
        query = query.offset((page - 1) * page_size).limit(page_size)

        result = await session.execute(query)
        leads = result.scalars().all()

        return LeadListResponse(
            total=total,
            items=[LeadResponse.model_validate(lead) for lead in leads],
            page=page,
            page_size=page_size
        )


def _apply_lead_filters(query, task_id, platform, intent_type, status, min_score, max_score, level, keyword, ip_location, current_user, start_ts=None, end_ts=None, role_tag=None):
    """通用筛选条件构造(给 list/export/stats 复用)

    重要: 日期范围(start_ts/end_ts,毫秒级)按【评论真实发布时间 create_time(秒级)】筛选,
    若 create_time 为空则 fallback 到 add_ts(入库时间,毫秒级)。用户预期是按评论发布时间过滤。
    """
    # 用户隔离过滤:admin 跨用户可见,普通用户只能看自己的
    if not is_admin(current_user):
        query = query.where(CustomerLead.owner_user_id == str(current_user["id"]))
    if task_id:
        query = query.where(CustomerLead.task_id == task_id)
    if platform:
        query = query.where(CustomerLead.platform == platform)
    if intent_type:
        query = query.where(CustomerLead.intent_type == intent_type)
    if status:
        query = query.where(CustomerLead.status == status)
    if level:
        lvl = level.lower()
        if lvl == "high":
            query = query.where(CustomerLead.lead_score >= 80)
        elif lvl == "medium":
            query = query.where(CustomerLead.lead_score >= 50).where(CustomerLead.lead_score < 80)
        elif lvl == "low":
            query = query.where(CustomerLead.lead_score < 50)
    if min_score is not None:
        query = query.where(CustomerLead.lead_score >= min_score)
    if max_score is not None:
        query = query.where(CustomerLead.lead_score <= max_score)
    if keyword:
        query = query.where(
            (CustomerLead.content.contains(keyword)) |
            (CustomerLead.title.contains(keyword)) |
            (CustomerLead.nickname.contains(keyword))
        )
    if ip_location:
        query = query.where(CustomerLead.ip_location.contains(ip_location))
    if role_tag:
        # 角色标签筛选:supplier供方/consumer需求方/neutral中性
        query = query.where(CustomerLead.role_tag == role_tag)
    # 日期范围筛选:优先用 create_time(评论发布时间,秒级),fallback 到 add_ts(入库时间,毫秒级)
    # start_ts/end_ts 来自前端,均为毫秒级
    if start_ts is not None:
        # create_time * 1000 将秒→毫秒,与 start_ts(毫秒)对齐;add_ts 已是毫秒
        query = query.where(
            func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) >= start_ts
        )
    if end_ts is not None:
        query = query.where(
            func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) <= end_ts
        )
    return query


async def _load_task_target_regions(session, task_id: Optional[str]) -> Optional[List[str]]:
    """加载任务的目标地区列表(用于自动地区过滤,与 list_leads/export 完全一致)。

    若任务设置了 target_regions,返回地区列表;否则返回 None。
    批量采集/监测接口必须复用此逻辑,否则会出现"列表已按地区过滤、但批量未过滤"的数量不一致问题。
    """
    if not task_id:
        return None
    try:
        from database.models import CrawlerTaskModel
        import json as _json
        result = await session.execute(
            select(CrawlerTaskModel.target_regions).where(CrawlerTaskModel.id == task_id)
        )
        row = result.first()
        if row and row[0]:
            regions = _json.loads(row[0])
            if regions and isinstance(regions, list) and len(regions) > 0:
                return regions
    except Exception:
        pass
    return None


def _apply_target_regions_filter(query, regions: Optional[List[str]], ip_location: Optional[str] = None):
    """应用任务自动地区过滤(与 list_leads 一致):仅当未单独指定 ip_location 时,按 target_regions 过滤。"""
    if regions and not ip_location:
        from sqlalchemy import or_
        conditions = [CustomerLead.ip_location.contains(r) for r in regions]
        if conditions:
            query = query.where(or_(*conditions))
    return query


@router.get("/export")
async def export_leads(
    task_id: Optional[str] = None,
    platform: Optional[str] = None,
    intent_type: Optional[str] = None,
    status: Optional[str] = None,
    min_score: Optional[int] = None,
    max_score: Optional[int] = None,
    level: Optional[str] = None,
    keyword: Optional[str] = None,
    ip_location: Optional[str] = None,
    start_ts: Optional[int] = None,
    end_ts: Optional[int] = None,
    role_tag: Optional[str] = None,
    format: str = "xlsx",
    current_user: dict = Depends(get_current_user),
):
    """导出获客线索为 Excel(.xlsx) 或 CSV(.csv)

    支持与 /leads/list 相同的筛选条件,导出全量数据(非分页)。
    Excel 包含完整字段:用户信息 + 评论 + 源视频 + 意向评分 + IP属地,
    CSV 为纯文本格式,无需任何解密/转换,直接用 Excel/WPS/记事本打开即可阅读。
    """
    import json
    from sqlalchemy import or_ as sql_or

    async with async_db_session() as session:
        auto_region_regions = None
        if task_id:
            try:
                from database.models import CrawlerTaskModel
                task_result = await session.execute(
                    select(CrawlerTaskModel.target_regions).where(CrawlerTaskModel.id == task_id)
                )
                task_row = task_result.first()
                if task_row and task_row[0]:
                    regions = json.loads(task_row[0])
                    if regions and isinstance(regions, list) and len(regions) > 0:
                        auto_region_regions = regions
            except Exception:
                pass

        query = select(CustomerLead)
        query = _apply_lead_filters(query, task_id, platform, intent_type, status,
                                     min_score, max_score, level, keyword, ip_location, current_user,
                                     start_ts, end_ts, role_tag)

        if auto_region_regions and not ip_location:
            conditions = []
            for region in auto_region_regions:
                conditions.append(CustomerLead.ip_location.contains(region))
            query = query.where(sql_or(*conditions))

        query = query.order_by(desc(CustomerLead.lead_score), desc(CustomerLead.add_ts))
        query = query.limit(10000)
        result = await session.execute(query)
        leads = result.scalars().all()

        task_name = ""
        if task_id:
            try:
                tr = await session.execute(select(CrawlerTaskModel.name).where(CrawlerTaskModel.id == task_id))
                row = tr.first()
                if row:
                    task_name = row[0] or ""
            except Exception:
                pass

    platform_map = {"douyin": "抖音", "xhs": "小红书", "kuaishou": "快手", "weibo": "微博", "bilibili": "B站"}
    intent_map = {"purchase": "购买意向", "cooperation": "合作意向", "inquiry": "咨询意向",
                  "potential": "潜在需求", "discussion": "一般关注"}
    status_map = {"new": "新线索", "contacted": "已联系", "qualified": "已确认", "converted": "已成交", "ignored": "已忽略"}
    home_url_map = {
        "douyin": "https://www.douyin.com/user/{sec_uid}",
        "xhs": "https://www.xiaohongshu.com/user/profile/{user_id}",
        "kuaishou": "https://www.kuaishou.com/profile/{user_id}",
        "weibo": "https://weibo.com/u/{user_id}",
        "bilibili": "https://space.bilibili.com/{user_id}",
    }
    role_map = {"supplier": "合作厂家", "consumer": "需求方", "neutral": "中性"}

    def _build_row(lead):
        level_cn = "高" if lead.lead_score >= 80 else ("中" if lead.lead_score >= 50 else "低")
        platform_cn = platform_map.get(lead.platform or "", lead.platform or "")
        intent_cn = intent_map.get(lead.intent_type or "", lead.intent_type or "")
        status_cn = status_map.get(lead.status or "", lead.status or "")
        role_cn = role_map.get(lead.role_tag or "", "中性")
        home_url = ""
        if lead.sec_uid and lead.platform in ("douyin",):
            home_url = f"https://www.douyin.com/user/{lead.sec_uid}"
        elif lead.user_id and lead.platform in ("xhs", "kuaishou", "weibo", "bilibili"):
            tpl = home_url_map.get(lead.platform, "")
            home_url = tpl.format(user_id=lead.user_id) if tpl else ""
        add_time = ""
        try:
            if lead.add_ts:
                add_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(lead.add_ts / 1000))
        except Exception:
            pass
        like_count = getattr(lead, 'like_count', 0) or 0
        source_desc = lead.source_video_desc or ""
        source_desc_short = source_desc[:100] + ("..." if len(source_desc) > 100 else "")
        return [
            level_cn, lead.lead_score or 0, intent_cn,
            lead.matched_keywords or "", lead.notes or "",
            lead.nickname or "", lead.user_id or "", lead.sec_uid or "", home_url,
            lead.ip_location or "", platform_cn,
            lead.content or "", like_count, add_time,
            lead.source_video_title or "", source_desc_short, lead.source_author_nickname or "", lead.source_video_url or "",
            lead.task_id or "", status_cn, role_cn,
        ]

    if format == "csv":
        import csv
        import io as _io
        headers = [
            "意向等级", "意向评分", "意向类型", "匹配关键词", "评分原因",
            "用户昵称", "用户ID", "SEC_UID", "用户主页",
            "IP属地", "平台",
            "评论内容", "点赞数", "评论时间",
            "源视频标题", "源视频描述", "源视频作者", "源视频链接",
            "任务ID", "状态", "角色",
        ]
        buf = _io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf)
        writer.writerow(headers)
        for lead in leads:
            writer.writerow(_build_row(lead))
        content = buf.getvalue().encode("utf-8")

        name_parts = []
        if task_name:
            safe_name = "".join(c for c in task_name if c not in '\\/:*?"<>|')[:30]
            name_parts.append(safe_name)
        if level:
            name_parts.append({"high": "高意向", "medium": "中意向", "low": "低意向"}.get(level.lower(), level))
        if ip_location:
            name_parts.append(ip_location)
        name_parts.append(f"{len(leads)}条")
        name_parts.append(time.strftime("%Y%m%d_%H%M"))
        filename = "_".join(name_parts) + ".csv"
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([content]),
            media_type="text/csv; charset=utf-8-sig",
            headers={
                "Content-Disposition": f"attachment; filename=\"leads.csv\"; filename*=UTF-8''{encoded_filename}",
            },
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "获客线索"

    headers = [
        "序号", "意向等级", "意向评分", "意向类型", "匹配关键词", "评分原因",
        "用户昵称", "用户ID", "SEC_UID", "用户主页",
        "IP属地", "平台",
        "评论内容", "点赞数", "评论时间",
        "源视频标题", "源视频描述", "源视频作者", "源视频链接",
        "任务ID", "状态", "角色",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for idx, lead in enumerate(leads, start=1):
        row_data = [idx] + _build_row(lead)
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=idx + 1, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if col_idx == 2:
                if val == "高":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif val == "中":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500", bold=True)

    col_widths = [6, 10, 10, 12, 24, 30, 18, 18, 22, 36, 12, 10, 50, 8, 18, 40, 40, 16, 40, 22, 10, 8]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    name_parts = []
    if task_name:
        safe_name = "".join(c for c in task_name if c not in '\\/:*?"<>|')[:30]
        name_parts.append(safe_name)
    if level:
        name_parts.append({"high": "高意向", "medium": "中意向", "low": "低意向"}.get(level.lower(), level))
    if ip_location:
        name_parts.append(ip_location)
    name_parts.append(f"{len(leads)}条")
    name_parts.append(time.strftime("%Y%m%d_%H%M"))
    filename = "_".join(name_parts) + ".xlsx"
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"leads.xlsx\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/regions")
async def get_lead_regions(
    task_id: Optional[str] = None,
    level: Optional[str] = None,
    limit: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user),
):
    """聚合统计当前用户线索库中的 IP 属地分布(Top N),用于前端快捷标签。

    返回示例:[{"ip_location": "四川", "count": 120}, {"ip_location": "广东", "count": 88}]
    空 IP 属地不计入。
    level 参数与 /leads/list 保持一致,避免地域标签数量与列表不一致。
    """
    async with async_db_session() as session:
        query = (
            select(CustomerLead.ip_location, func.count().label("cnt"))
            .where(CustomerLead.ip_location != None)
            .where(CustomerLead.ip_location != "")
        )
        cond = user_scope_filter(current_user, CustomerLead)
        if cond is not None:
            query = query.where(cond)
        if task_id:
            query = query.where(CustomerLead.task_id == task_id)
        if level:
            lvl = level.lower()
            if lvl == "high":
                query = query.where(CustomerLead.lead_score >= 80)
            elif lvl == "medium":
                query = query.where(CustomerLead.lead_score >= 50).where(CustomerLead.lead_score < 80)
            elif lvl == "low":
                query = query.where(CustomerLead.lead_score < 50)
        query = query.group_by(CustomerLead.ip_location).order_by(desc("cnt")).limit(limit)
        result = await session.execute(query)
        return [{"ip_location": r[0], "count": r[1]} for r in result]


@router.get("/stats", response_model=LeadStatsResponse)
async def get_lead_stats(
    task_id: Optional[str] = None,
    keyword: Optional[str] = None,
    level: Optional[str] = None,
    ip_location: Optional[str] = None,
    platform: Optional[str] = None,
    role_tag: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """获取获客线索统计信息(按用户隔离)

    统计维度不传 status/start_ts/end_ts(status 是要统计的维度,时间范围不影响全局统计)。
    keyword 筛选与 list/export 保持一致(匹配 content/title/nickname),
    避免前端"统计总数 ≠ 列表 total"的体验问题。
    """
    import json
    from sqlalchemy import or_ as sql_or
    async with async_db_session() as session:
        # 自动获取任务的 target_regions 并应用地区过滤
        auto_region_regions = None
        if task_id:
            try:
                from database.models import CrawlerTaskModel
                task_result = await session.execute(
                    select(CrawlerTaskModel.target_regions).where(CrawlerTaskModel.id == task_id)
                )
                task_row = task_result.first()
                if task_row and task_row[0]:
                    regions = json.loads(task_row[0])
                    if regions and isinstance(regions, list) and len(regions) > 0:
                        auto_region_regions = regions
            except Exception:
                pass

        # 基础查询(复用通用筛选,与 list/export 保持一致)
        base_query = select(CustomerLead)
        base_query = _apply_lead_filters(
            base_query, task_id, platform, None, None,
            None, None, level, keyword, ip_location,
            current_user, None, None, role_tag,
        )

        # 应用任务自动地区过滤(如果有设置 target_regions)
        if auto_region_regions and not ip_location:
            conditions = []
            for region in auto_region_regions:
                conditions.append(CustomerLead.ip_location.contains(region))
            base_query = base_query.where(sql_or(*conditions))
        
        # 总数
        total_result = await session.execute(
            select(func.count()).select_from(base_query.subquery())
        )
        total_leads = total_result.scalar()
        
        # 各状态数量
        status_counts = {}
        for status in ["new", "pending", "contacted", "qualified", "converted", "failed", "ignored"]:
            count_query = select(func.count()).select_from(
                base_query.where(CustomerLead.status == status).subquery()
            )
            result = await session.execute(count_query)
            status_counts[status] = result.scalar()
        
        # 平台分布
        sub = base_query.subquery()
        platform_result = await session.execute(
            select(sub.c.platform, func.count())
            .group_by(sub.c.platform)
        )
        platform_distribution = {row[0]: row[1] for row in platform_result.all()}
        
        # 意图分布
        intent_result = await session.execute(
            select(sub.c.intent_type, func.count())
            .group_by(sub.c.intent_type)
        )
        intent_distribution = {row[0]: row[1] for row in intent_result.all()}

        # 角色分布(supplier供方/consumer需求方/neutral中性)
        role_result = await session.execute(
            select(sub.c.role_tag, func.count())
            .group_by(sub.c.role_tag)
        )
        role_distribution = {"supplier": 0, "consumer": 0, "neutral": 0}
        for row in role_result.all():
            tag = row[0] or "neutral"
            role_distribution[tag] = row[1]

        # 平均评分
        avg_result = await session.execute(
            select(func.avg(sub.c.lead_score))
        )
        avg_score = avg_result.scalar() or 0

        return LeadStatsResponse(
            total_leads=total_leads,
            new_leads=status_counts.get("new", 0),
            pending_leads=status_counts.get("pending", 0),
            contacted_leads=status_counts.get("contacted", 0),
            qualified_leads=status_counts.get("qualified", 0),
            converted_leads=status_counts.get("converted", 0),
            failed_leads=status_counts.get("failed", 0),
            ignored_leads=status_counts.get("ignored", 0),
            platform_distribution=platform_distribution,
            intent_distribution=intent_distribution,
            avg_lead_score=round(float(avg_score), 2),
            role_distribution=role_distribution
        )


@router.post("/{lead_id}/status")
async def update_lead_status(
    lead_id: int, 
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """更新线索状态"""
    from sqlalchemy import update
    import time
    
    status = data.get("status")
    notes = data.get("notes")
    
    if not status:
        raise HTTPException(status_code=400, detail="status is required")

    async with async_db_session() as session:
        # 先校验归属
        result = await session.execute(select(CustomerLead).where(CustomerLead.id == lead_id))
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        if not is_admin(current_user):
            owner = getattr(lead, "owner_user_id", "") or ""
            if owner and owner != str(current_user["id"]):
                raise HTTPException(status_code=403, detail="无权操作该线索")

        update_data = {"status": status, "last_modify_ts": int(time.time() * 1000)}
        if notes:
            update_data["notes"] = notes

        stmt = update(CustomerLead).where(CustomerLead.id == lead_id).values(**update_data)
        await session.execute(stmt)
        await session.commit()

        return {"success": True, "message": f"Lead {lead_id} status updated to {status}"}


@router.get("/{lead_id}", response_model=LeadResponse)
async def get_lead_detail(lead_id: int, current_user: dict = Depends(get_current_user)):
    """获取线索详情(含来源视频信息,用于营销时查看原视频/原文案)"""
    from database.models import DouyinAweme, DouyinAwemeComment

    async with async_db_session() as session:
        result = await session.execute(
            select(CustomerLead).where(CustomerLead.id == lead_id)
        )
        lead = result.scalar_one_or_none()

        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        if not is_admin(current_user):
            owner = getattr(lead, "owner_user_id", "") or ""
            if owner and owner != str(current_user["id"]):
                raise HTTPException(status_code=403, detail="无权访问该线索")

        lead_dict = {
            "id": lead.id, "task_id": lead.task_id, "platform": lead.platform,
            "data_type": lead.data_type, "data_id": lead.data_id,
            "user_id": lead.user_id, "sec_uid": lead.sec_uid,
            "nickname": lead.nickname, "avatar": lead.avatar,
            "ip_location": lead.ip_location, "content": lead.content,
            "title": lead.title, "url": lead.url,
            "matched_keywords": lead.matched_keywords,
            "intent_type": lead.intent_type, "lead_score": lead.lead_score,
            "status": lead.status, "notes": lead.notes,
            "add_ts": lead.add_ts, "last_modify_ts": lead.last_modify_ts,
            "create_time": getattr(lead, "create_time", None),
            "source_aweme_id": "",
            "source_video_title": "",
            "source_video_desc": "",
            "source_video_url": "",
            "source_cover_url": getattr(lead, "source_cover_url", "") or "",
            "source_author_nickname": getattr(lead, "source_author_nickname", "") or "",
            # 增强字段(客户需求:支持复制和打开链接)
            "comment_url": getattr(lead, "comment_url", "") or "",
            "profile_url": getattr(lead, "profile_url", "") or "",
            "platform_display_id": getattr(lead, "platform_display_id", "") or "",
            # 用户主页联系方式(一键获客时从用户主页采集)
            "contact_phone": getattr(lead, "contact_phone", "") or "",
            "contact_wechat": getattr(lead, "contact_wechat", "") or "",
            "bio_text": getattr(lead, "bio_text", "") or "",
            "contact_status": getattr(lead, "contact_status", "pending") or "pending",
            "reply_monitor_ts": getattr(lead, "reply_monitor_ts", 0) or 0,
        }

        # 评论类型线索:通过 comment_id → aweme_id → DouyinAweme 关联原视频
        data_id = lead.data_id or ""
        if lead.platform == "douyin" and data_id:
            aweme_id = ""
            # 1. 从评论表查 aweme_id
            try:
                cmt = await session.execute(
                    select(DouyinAwemeComment.aweme_id)
                    .where(DouyinAwemeComment.comment_id == data_id)
                    .limit(1)
                )
                aweme_id = cmt.scalar() or ""
            except Exception:
                pass
            # 2. 如果线索本身就是视频(data_type=video),data_id 就是 aweme_id
            if not aweme_id and lead.data_type == "video":
                aweme_id = data_id
            # 3. 查 DouyinAweme 获取视频信息
            if aweme_id:
                try:
                    awe = await session.execute(
                        select(DouyinAweme)
                        .where(DouyinAweme.aweme_id == aweme_id)
                        .limit(1)
                    )
                    aweme = awe.scalar_one_or_none()
                    if aweme:
                        lead_dict["source_aweme_id"] = aweme.aweme_id or ""
                        lead_dict["source_video_title"] = aweme.title or ""
                        lead_dict["source_video_desc"] = aweme.desc or ""
                        lead_dict["source_video_url"] = aweme.aweme_url or ""
                        lead_dict["source_cover_url"] = aweme.cover_url or ""
                        lead_dict["source_author_nickname"] = aweme.nickname or ""
                except Exception:
                    pass

        return LeadResponse(**lead_dict)


@router.delete("/{lead_id}")
async def delete_lead(lead_id: int, current_user: dict = Depends(get_current_user)):
    """删除单条线索"""
    from sqlalchemy import delete as sql_delete

    async with async_db_session() as session:
        result = await session.execute(
            select(CustomerLead).where(CustomerLead.id == lead_id)
        )
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        if not is_admin(current_user):
            owner = getattr(lead, "owner_user_id", "") or ""
            if owner and owner != str(current_user["id"]):
                raise HTTPException(status_code=403, detail="无权操作该线索")

        await session.execute(sql_delete(CustomerLead).where(CustomerLead.id == lead_id))
        await session.commit()
        return {"success": True, "message": f"Lead {lead_id} deleted"}


class BatchDeleteRequest(BaseModel):
    ids: List[int]


@router.post("/batch-delete")
async def batch_delete_leads(request: BatchDeleteRequest, current_user: dict = Depends(get_current_user)):
    """批量删除线索(仅删除当前用户拥有的,管理员删除所有)"""
    from sqlalchemy import delete as sql_delete

    if not request.ids:
        raise HTTPException(status_code=400, detail="No ids provided")

    async with async_db_session() as session:
        # 非管理员:只能删除自己的线索
        if is_admin(current_user):
            result = await session.execute(
                sql_delete(CustomerLead).where(CustomerLead.id.in_(request.ids))
            )
        else:
            result = await session.execute(
                sql_delete(CustomerLead).where(
                    CustomerLead.id.in_(request.ids),
                    CustomerLead.owner_user_id == str(current_user["id"]),
                )
            )
        await session.commit()
        return {"success": True, "deleted_count": result.rowcount}


@router.delete("/task/{task_id}")
async def delete_leads_by_task(task_id: str, current_user: dict = Depends(get_current_user)):
    """删除指定任务的所有线索"""
    from sqlalchemy import delete as sql_delete

    async with async_db_session() as session:
        if is_admin(current_user):
            result = await session.execute(
                sql_delete(CustomerLead).where(CustomerLead.task_id == task_id)
            )
        else:
            result = await session.execute(
                sql_delete(CustomerLead).where(
                    CustomerLead.task_id == task_id,
                    CustomerLead.owner_user_id == str(current_user["id"]),
                )
            )
        await session.commit()
        return {"success": True, "deleted_count": result.rowcount}


@router.post("/import-file")
async def import_leads_file(
    task_id: str = Query(..., description="归属任务ID(必填)"),
    platform: str = Query("manual", description="平台标识,默认 manual"),
    file: UploadFile = File(..., description="CSV/Excel 文件"),
    current_user: dict = Depends(get_current_user),
):
    """批量导入线索(文件上传版)- 支持 CSV/Excel

    必填列: content
    可选列: nickname, user_id, ip_location, intent_type, lead_score, status, notes, url
    """
    import csv
    import io
    from sqlalchemy import insert

    owner_uid = str(current_user["id"])
    filename = (file.filename or "").lower()

    async def _parse_csv(raw: bytes) -> List[dict]:
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    async def _parse_excel(raw: bytes) -> List[dict]:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="服务器未安装 openpyxl,无法解析 Excel")
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            result.append({headers[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(headers))})
        return result

    raw = await file.read()
    if filename.endswith(".csv"):
        records = await _parse_csv(raw)
    elif filename.endswith((".xlsx", ".xls")):
        records = await _parse_excel(raw)
    else:
        raise HTTPException(status_code=400, detail="仅支持 .csv / .xlsx / .xls 格式")

    if not records:
        raise HTTPException(status_code=400, detail="文件无有效数据行")

    # 字段映射 + 校验
    now_ms = int(time.time() * 1000)
    batch: List[dict] = []
    skipped = 0
    for r in records:
        content = (r.get("content") or r.get("咨询内容") or "").strip()
        if not content:
            skipped += 1
            continue
        def _get(*keys: str) -> str:
            for k in keys:
                v = r.get(k)
                if v is not None and str(v).strip():
                    return str(v).strip()
            return ""
        try:
            score = int(float(r.get("lead_score") or r.get("线索评分") or 0))
        except (ValueError, TypeError):
            score = 0
        batch.append({
            "task_id": task_id,
            "platform": platform,
            "data_type": "import",
            "data_id": "",
            "user_id": _get("user_id", "用户ID"),
            "sec_uid": "",
            "nickname": _get("nickname", "昵称"),
            "avatar": "",
            "ip_location": _get("ip_location", "IP属地"),
            "content": content,
            "title": "",
            "url": _get("url", "链接"),
            "matched_keywords": "",
            "intent_type": _get("intent_type", "意向类型") or "discussion",
            "lead_score": max(0, min(100, score)),
            "status": _get("status", "状态") or "new",
            "notes": _get("notes", "备注"),
            "add_ts": now_ms,
            "last_modify_ts": now_ms,
            "create_time": None,
            "owner_user_id": owner_uid,
        })

    if not batch:
        raise HTTPException(status_code=400, detail=f"无有效线索(全部缺少 content 列,跳过 {skipped} 行)")

    # 批量插入(每 500 条一批)
    BATCH = 500
    inserted = 0
    async with async_db_session() as session:
        for i in range(0, len(batch), BATCH):
            chunk = batch[i:i + BATCH]
            await session.execute(insert(CustomerLead), chunk)
            await session.commit()
            inserted += len(chunk)

    # 推送新线索事件(复用 WebSocket 通道)
    try:
        from .websocket import notify_new_leads
        await notify_new_leads(owner_uid, task_id, platform, inserted, 0, 0, 0)
    except Exception:
        pass

    return {
        "success": True,
        "imported": inserted,
        "skipped": skipped,
        "total_rows": len(records),
    }


# ==================== 联系方式采集 + 评论回复监测 ====================
class BatchContactRequest(BaseModel):
    lead_ids: List[int] = []
    task_id: Optional[str] = None
    # 可选: 按筛选条件批量采集(与列表筛选一致)
    platform: Optional[str] = None
    role_tag: Optional[str] = None
    ip_location: Optional[str] = None
    status: Optional[str] = None
    # 意向等级筛选(high>=80 / medium 50-79 / low<50),与列表查询 level 参数一致
    level: Optional[str] = None
    # 日期范围筛选(毫秒级,与列表筛选一致,优先用 create_time 评论发布时间)
    start_ts: Optional[int] = None
    end_ts: Optional[int] = None
    limit: int = 100  # 防止一次采集过多触发风控


@router.post("/{lead_id}/collect-contact")
async def collect_lead_contact(lead_id: int, current_user: dict = Depends(get_current_user)):
    """采集单条线索的用户主页联系方式(手机号/微信号/简介)。"""
    from ..services.contact_collector import collect_contact_for_lead
    owner_uid = str(current_user["id"])
    result = await collect_contact_for_lead(lead_id, owner_uid)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "采集失败"))
    return result


@router.post("/collect-contacts-batch")
async def collect_contacts_batch(request: BatchContactRequest, current_user: dict = Depends(get_current_user)):
    """批量采集联系方式: 支持 lead_ids 直选 或 按 task_id+筛选条件批量。"""
    from ..services.contact_collector import collect_contacts_batch as _collect
    owner_uid = str(current_user["id"])

    lead_ids = list(request.lead_ids or [])
    # 若未指定 lead_ids,按筛选条件查询(与列表筛选一致)
    if not lead_ids and request.task_id:
        async with async_db_session() as session:
            if session is None:
                raise HTTPException(status_code=500, detail="数据库不可用")
            q = select(CustomerLead.id).where(
                CustomerLead.task_id == request.task_id,
                CustomerLead.owner_user_id == owner_uid,
                CustomerLead.sec_uid != "",
            )
            if request.platform:
                q = q.where(CustomerLead.platform == request.platform)
            if request.status:
                q = q.where(CustomerLead.status == request.status)
            if request.role_tag:
                q = q.where(CustomerLead.role_tag == request.role_tag)
            if request.ip_location:
                q = q.where(CustomerLead.ip_location.contains(request.ip_location))
            # 意向等级筛选:与列表 _apply_lead_filters 的 level 逻辑完全一致
            if request.level:
                lvl = request.level.lower()
                if lvl == "high":
                    q = q.where(CustomerLead.lead_score >= 80)
                elif lvl == "medium":
                    q = q.where(CustomerLead.lead_score >= 50).where(CustomerLead.lead_score < 80)
                elif lvl == "low":
                    q = q.where(CustomerLead.lead_score < 50)
            # 日期范围筛选:与列表一致,优先用 create_time(评论发布时间,秒级),fallback 到 add_ts(毫秒级)
            if request.start_ts is not None:
                q = q.where(
                    func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) >= request.start_ts
                )
            if request.end_ts is not None:
                q = q.where(
                    func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) <= request.end_ts
                )
            # 任务自动地区过滤(与 list_leads 完全一致):若未单独指定 ip_location,则按 target_regions 过滤
            # 修复:此前批量采集未应用此过滤,导致"列表显示1条但批量采集31条"的数量不一致
            auto_regions = await _load_task_target_regions(session, request.task_id)
            q = _apply_target_regions_filter(q, auto_regions, request.ip_location)
            q = q.limit(request.limit)
            lead_ids = (await session.execute(q)).scalars().all()

    if not lead_ids:
        raise HTTPException(status_code=400, detail="未找到符合条件的线索(需有 sec_uid)")

    # 注册批量任务,返回 job_id 供前端轮询进度
    job_id = f"contact_collect_{uuid.uuid4().hex[:8]}"
    _register_batch_job(job_id, request.task_id or "", "contact_collect", len(lead_ids))

    async def _run_with_progress():
        """包装 collect_contacts_batch,在结束后更新 job 状态。"""
        try:
            result = await _collect(lead_ids, owner_uid, job_id=job_id)
            _finish_batch_job(
                job_id,
                "completed" if result.get("success") else "failed",
                result.get("message", "采集完成"),
            )
        except Exception as e:
            _finish_batch_job(job_id, "failed", f"采集异常: {e}")

    asyncio.create_task(_run_with_progress())
    return {
        "success": True,
        "message": f"已开始后台采集 {len(lead_ids)} 条线索的联系方式",
        "job_id": job_id,
        "total": len(lead_ids),
    }


@router.post("/{lead_id}/monitor-replies")
async def monitor_lead_replies_ep(lead_id: int, current_user: dict = Depends(get_current_user)):
    """手动触发单条线索的评论回复监测。"""
    from ..services.comment_reply_monitor import monitor_lead_replies
    owner_uid = str(current_user["id"])
    result = await monitor_lead_replies(lead_id, owner_uid)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "监测失败"))
    return result


@router.post("/task/{task_id}/monitor-replies")
async def monitor_task_replies_ep(
    task_id: str,
    filters: dict = None,
    current_user: dict = Depends(get_current_user),
):
    """手动触发整个任务的评论回复监测(支持与列表一致的筛选条件)。"""
    from ..services.comment_reply_monitor import monitor_task_replies
    owner_uid = str(current_user["id"])
    # 兼容 filters 为 None 或非 dict 的情况
    if not isinstance(filters, dict):
        filters = {}

    # 与 monitor_task_replies 内部 limit 保持一致(默认 50),否则进度条 total 与实际处理量不符
    monitor_limit = int(filters.get("limit") or 50)

    # 预查询可监测线索数(受 monitor_limit 上限),用于注册 job 进度
    # 应用与列表一致的筛选条件(平台/角色/地域/日期),让 total 与前端筛选结果一致
    async with async_db_session() as session:
        total = 0
        if session is not None:
            q = select(func.count()).select_from(CustomerLead).where(
                CustomerLead.task_id == task_id,
                CustomerLead.owner_user_id == owner_uid,
                CustomerLead.source_aweme_id != "",
                CustomerLead.data_id != "",
            )
            if filters.get("platform"):
                q = q.where(CustomerLead.platform == filters["platform"])
            if filters.get("role_tag"):
                q = q.where(CustomerLead.role_tag == filters["role_tag"])
            if filters.get("ip_location"):
                q = q.where(CustomerLead.ip_location.contains(filters["ip_location"]))
            if filters.get("status"):
                q = q.where(CustomerLead.status == filters["status"])
            # 意向等级筛选:与列表 _apply_lead_filters 的 level 逻辑完全一致
            lvl = (filters.get("level") or "").lower() if filters.get("level") else ""
            if lvl == "high":
                q = q.where(CustomerLead.lead_score >= 80)
            elif lvl == "medium":
                q = q.where(CustomerLead.lead_score >= 50).where(CustomerLead.lead_score < 80)
            elif lvl == "low":
                q = q.where(CustomerLead.lead_score < 50)
            if filters.get("start_ts") is not None:
                q = q.where(
                    func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) >= int(filters["start_ts"])
                )
            if filters.get("end_ts") is not None:
                q = q.where(
                    func.coalesce(CustomerLead.create_time * 1000, CustomerLead.add_ts) <= int(filters["end_ts"])
                )
            # 任务自动地区过滤(与 list_leads 完全一致):若未单独指定 ip_location,则按 target_regions 过滤
            # 修复:此前监测任务未应用此过滤,导致 total 与列表筛选结果不一致
            auto_regions = await _load_task_target_regions(session, task_id)
            q = _apply_target_regions_filter(q, auto_regions, filters.get("ip_location"))
            cnt = await session.execute(q)
            full_count = cnt.scalar() or 0
            total = min(full_count, monitor_limit)

    if total == 0:
        return {"success": False, "message": "当前筛选条件下无可监测线索(需有源视频+评论ID)"}

    job_id = f"reply_monitor_{uuid.uuid4().hex[:8]}"
    _register_batch_job(job_id, task_id, "reply_monitor", total)

    async def _run_with_progress():
        try:
            result = await monitor_task_replies(task_id, owner_uid, limit=monitor_limit, job_id=job_id, filters=filters)
            _finish_batch_job(
                job_id,
                "completed" if result.get("success") else "failed",
                result.get("message", "监测完成"),
            )
        except Exception as e:
            _finish_batch_job(job_id, "failed", f"监测异常: {e}")

    asyncio.create_task(_run_with_progress())
    return {
        "success": True,
        "message": f"已开始后台监测任务 {task_id} 下线索的评论回复",
        "job_id": job_id,
        "total": total,
    }


@router.get("/batch-jobs/{job_id}")
async def get_batch_job_status(job_id: str, current_user: dict = Depends(get_current_user)):
    """获取批量任务状态(供前端轮询进度)。"""
    job = _batch_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    return job


@router.get("/{lead_id}/replies")
async def list_lead_replies(
    lead_id: int,
    limit: int = 100,
    offset: int = 0,
    current_user: dict = Depends(get_current_user),
):
    """查看某条线索监测到的评论回复列表。"""
    from database.models import LeadCommentReply
    async with async_db_session() as session:
        if session is None:
            return {"items": [], "total": 0}
        # 校验归属
        lead = (await session.execute(
            select(CustomerLead).where(CustomerLead.id == lead_id)
        )).scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="线索不存在")
        if not is_admin(current_user):
            owner = lead.owner_user_id or ""
            if owner and owner != str(current_user["id"]):
                raise HTTPException(status_code=403, detail="无权访问")

        q = select(LeadCommentReply).where(LeadCommentReply.lead_id == lead_id)
        cq = select(func.count()).select_from(LeadCommentReply).where(LeadCommentReply.lead_id == lead_id)
        total = (await session.execute(cq)).scalar() or 0
        result = await session.execute(q.order_by(desc(LeadCommentReply.create_time)).offset(offset).limit(limit))
        items = [
            {
                "id": r.id,
                "comment_id": r.comment_id or "",
                "parent_comment_id": r.parent_comment_id or "",
                "user_id": r.user_id or "",
                "nickname": r.nickname or "",
                "avatar": r.avatar or "",
                "content": r.content or "",
                "like_count": r.like_count or "0",
                "create_time": r.create_time or 0,
                "is_from_lead": r.is_from_lead or 0,
                "is_read": r.is_read or 0,
                "add_ts": r.add_ts or 0,
            }
            for r in result.scalars().all()
        ]
        return {"items": items, "total": total, "offset": offset, "limit": limit}


@router.post("/{lead_id}/replies/mark-read")
async def mark_lead_replies_read(lead_id: int, current_user: dict = Depends(get_current_user)):
    """将某条线索的评论回复标记为已读。"""
    from database.models import LeadCommentReply
    async with async_db_session() as session:
        if session is None:
            raise HTTPException(status_code=500, detail="数据库不可用")
        await session.execute(
            update(LeadCommentReply).where(LeadCommentReply.lead_id == lead_id).values(is_read=1)
        )
        await session.commit()
    return {"success": True, "message": "已标记为已读"}

