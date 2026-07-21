# -*- coding: utf-8 -*-
"""
X Twitter 工作台 - 数据导出路由

提供 CSV 和 Excel 格式的数据导出:
- GET /x-workbench/export/sent-comments?format=csv|xlsx  导出已发评论
- GET /x-workbench/export/replies?format=csv|xlsx        导出收到的回复
- GET /x-workbench/export/analytics?format=xlsx          导出效果分析(多 sheet)

限制:
- 单次最多导出 export_max_rows 条(默认 10000),防止 OOM
- 导出前需要认证(与工作台其他接口一致)
"""
import csv
import io
import logging
import time
from typing import Any, Dict, List

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, and_, desc

from database.db_session import get_session
from database.models import XTwitterSentComment, XTwitterReply
from api.services.auth import get_current_user
from api.utils.rate_limit import rate_limit
from api.utils.workbench_config import workbench_config


router = APIRouter(
    prefix="/x-workbench/export",
    tags=["x-twitter-workbench"],
    dependencies=[
        Depends(get_current_user),
        Depends(rate_limit()),
    ],
)

logger = logging.getLogger("x_workbench_export")


# ==================== 工具函数 ====================

def _make_csv_response(rows: List[Dict[str, Any]], filename: str) -> StreamingResponse:
    """把 dict 列表转为 CSV StreamingResponse(带 BOM,Excel 可正确识别中文)"""
    output = io.StringIO()
    # 写入 UTF-8 BOM,确保 Excel 打开时中文不乱码
    output.write("\ufeff")
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    content = output.getvalue().encode("utf-8")
    output.close()
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}.csv"',
        "Content-Type": "text/csv; charset=utf-8",
    }
    return StreamingResponse(io.BytesIO(content), media_type="text/csv", headers=headers)


def _make_xlsx_response(sheets: Dict[str, List[Dict[str, Any]]], filename: str) -> StreamingResponse:
    """把多 sheet 数据转为 Excel StreamingResponse"""
    from openpyxl import Workbook

    wb = Workbook()
    # 移除默认 sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        # Excel sheet 名最长 31 字符
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            ws.append(["(无数据)"])
            continue
        # 写表头
        headers = list(rows[0].keys())
        ws.append(headers)
        # 加粗表头
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # 写数据
        for row in rows:
            ws.append([_to_excel_value(row.get(h)) for h in headers])
        # 自动列宽(粗略)
        for i, h in enumerate(headers, 1):
            max_len = max(len(str(h)), max((len(str(row.get(h, ""))) for row in rows), default=0))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _to_excel_value(v: Any) -> Any:
    """转换单元格值(时间戳转可读时间,None 转空串)"""
    if v is None:
        return ""
    if isinstance(v, (int, float)) and v > 1_000_000_000 and v < 10_000_000_000:
        # 看起来是时间戳(秒级),转为可读时间
        try:
            return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(v)))
        except Exception:
            return v
    return v


def _sent_comment_to_row(c: XTwitterSentComment) -> dict:
    """已发评论转导出行"""
    return {
        "ID": c.id,
        "推文ID": c.post_id,
        "推文URL": c.post_url,
        "推文作者": c.post_username,
        "推文内容": (c.post_content or "")[:200],
        "评论内容": c.comment_content,
        "评论URL": c.comment_url,
        "发送状态": c.sent_status,
        "失败原因": c.sent_error,
        "发送时间": c.sent_at,
        "来源": c.source,
        "监控中": "是" if c.monitoring == 1 else "否",
        "回复数": c.reply_count or 0,
        "AI已回复数": c.auto_replied_count or 0,
        "创建时间": c.add_ts,
    }


def _reply_to_row(r: XTwitterReply, sc_map: Dict[int, XTwitterSentComment] = None) -> dict:
    """回复转导出行"""
    sc = (sc_map or {}).get(r.sent_comment_id)
    return {
        "回复ID": r.id,
        "父评论ID": r.sent_comment_id,
        "父评论内容": (sc.comment_content if sc else "")[:100],
        "推文ID": r.post_id,
        "回复者用户名": r.replier_username,
        "回复者昵称": r.replier_nickname,
        "回复内容": r.reply_content,
        "回复URL": r.reply_url,
        "点赞数": r.reply_likes_count,
        "回复时间": r.reply_created_at,
        "AI回复状态": r.auto_reply_status,
        "AI回复内容": r.auto_reply_content,
        "AI回复URL": r.auto_reply_url,
        "AI回复时间": r.auto_replied_at,
        "入库时间": r.add_ts,
    }


# ==================== 端点 ====================

@router.get("/sent-comments")
async def export_sent_comments(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    status: str = Query("", description="按状态筛选: success/failed/draft"),
    start_ts: int = Query(0),
    end_ts: int = Query(0),
):
    """导出已发评论

    支持按状态和时间范围筛选,最多导出 export_max_rows 条。
    """
    max_rows = workbench_config.export_max_rows
    conditions = []
    if status:
        conditions.append(XTwitterSentComment.sent_status == status)
    if start_ts > 0:
        conditions.append(XTwitterSentComment.sent_at >= start_ts)
    if end_ts > 0:
        conditions.append(XTwitterSentComment.sent_at <= end_ts)

    async with get_session() as session:
        stmt = (
            select(XTwitterSentComment)
            .order_by(desc(XTwitterSentComment.id))
            .limit(max_rows)
        )
        if conditions:
            stmt = stmt.where(and_(*conditions))
        result = await session.execute(stmt)
        comments = result.scalars().all()

    rows = [_sent_comment_to_row(c) for c in comments]
    logger.info(f"导出 {len(rows)} 条已发评论(format={format})")

    ts = time.strftime("%Y%m%d_%H%M%S")
    if format.lower() == "xlsx":
        return _make_xlsx_response({"已发评论": rows}, f"sent_comments_{ts}")
    return _make_csv_response(rows, f"sent_comments_{ts}")


@router.get("/replies")
async def export_replies(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    status: str = Query("", description="按AI回复状态筛选: pending/sent/failed"),
    start_ts: int = Query(0),
    end_ts: int = Query(0),
):
    """导出收到的回复

    包含回复内容及 AI 自动回复情况。
    """
    max_rows = workbench_config.export_max_rows
    conditions = []
    if status:
        conditions.append(XTwitterReply.auto_reply_status == status)
    if start_ts > 0:
        conditions.append(XTwitterReply.add_ts >= start_ts)
    if end_ts > 0:
        conditions.append(XTwitterReply.add_ts <= end_ts)

    async with get_session() as session:
        # 查回复
        stmt = (
            select(XTwitterReply)
            .order_by(desc(XTwitterReply.id))
            .limit(max_rows)
        )
        if conditions:
            stmt = stmt.where(and_(*conditions))
        result = await session.execute(stmt)
        replies = result.scalars().all()

        # 批量查询父评论(用于展示父评论内容)
        sc_ids = {r.sent_comment_id for r in replies if r.sent_comment_id}
        sc_map: Dict[int, XTwitterSentComment] = {}
        if sc_ids:
            sc_result = await session.execute(
                select(XTwitterSentComment).where(XTwitterSentComment.id.in_(list(sc_ids)))
            )
            for sc in sc_result.scalars().all():
                sc_map[sc.id] = sc

    rows = [_reply_to_row(r, sc_map) for r in replies]
    logger.info(f"导出 {len(rows)} 条回复(format={format})")

    ts = time.strftime("%Y%m%d_%H%M%S")
    if format.lower() == "xlsx":
        return _make_xlsx_response({"回复记录": rows}, f"replies_{ts}")
    return _make_csv_response(rows, f"replies_{ts}")


@router.get("/analytics")
async def export_analytics(format: str = Query("xlsx", description="导出格式: xlsx(多 sheet)")):
    """导出完整效果分析报告(Excel 多 sheet)

    包含:
    - Sheet1: 已发评论(含回复数、AI回复数)
    - Sheet2: 回复记录(含 AI 回复情况)
    - Sheet3: 按作者分组统计
    """
    max_rows = workbench_config.export_max_rows

    async with get_session() as session:
        # 已发评论
        sc_result = await session.execute(
            select(XTwitterSentComment)
            .order_by(desc(XTwitterSentComment.id))
            .limit(max_rows)
        )
        comments = sc_result.scalars().all()
        sc_rows = [_sent_comment_to_row(c) for c in comments]

        # 回复记录
        reply_result = await session.execute(
            select(XTwitterReply)
            .order_by(desc(XTwitterReply.id))
            .limit(max_rows)
        )
        replies = reply_result.scalars().all()

        # 父评论映射
        sc_ids = {r.sent_comment_id for r in replies if r.sent_comment_id}
        sc_map: Dict[int, XTwitterSentComment] = {}
        if sc_ids:
            sc_result2 = await session.execute(
                select(XTwitterSentComment).where(XTwitterSentComment.id.in_(list(sc_ids)))
            )
            for sc in sc_result2.scalars().all():
                sc_map[sc.id] = sc
        reply_rows = [_reply_to_row(r, sc_map) for r in replies]

        # 按作者分组统计
        from sqlalchemy import func, case
        topic_result = await session.execute(
            select(
                XTwitterSentComment.post_username,
                func.count(XTwitterSentComment.id).label("total"),
                func.sum(case(
                    (XTwitterSentComment.sent_status == "success", 1),
                    else_=0,
                )).label("success_count"),
                func.sum(XTwitterSentComment.reply_count).label("reply_count"),
                func.sum(XTwitterSentComment.auto_replied_count).label("ai_replied"),
            )
            .group_by(XTwitterSentComment.post_username)
            .order_by(desc(func.sum(XTwitterSentComment.reply_count)))
        )
        topic_rows = [
            {
                "作者": row[0] or "(未知)",
                "评论数": row[1] or 0,
                "成功数": row[2] or 0,
                "回复数": row[3] or 0,
                "AI回复数": row[4] or 0,
                "回复率%": round((row[3] or 0) / max(1, row[2] or 0) * 100, 1),
            }
            for row in topic_result.all()
        ]

    sheets = {
        "已发评论": sc_rows,
        "回复记录": reply_rows,
        "按作者统计": topic_rows,
    }
    logger.info(f"导出效果分析报告: {len(sc_rows)} 评论 + {len(reply_rows)} 回复 + {len(topic_rows)} 作者")

    ts = time.strftime("%Y%m%d_%H%M%S")
    return _make_xlsx_response(sheets, f"analytics_report_{ts}")
